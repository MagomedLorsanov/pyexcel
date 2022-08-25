from openpyxl import Workbook, load_workbook

def CreateWorkbook(workbook_path):    
    workbook = Workbook()
    workbook.save(workbook_path)
    return workbook_path

def CreateSheet(workbook, sheet_name):
    wb.create_sheet(sheet_name)
    return sheet_name

def FillRow(char,numofchars,value):
    cells= f'{char}{numofchars}'
    sheet[cells].value= value

ItogoSerch = 'ИТОГО:'
kolledgeRating = {}
kelledgeInfo = {}
kollegePoints = []
kolledgeSortedRating = {}
abat = ['','B','C','D','E','F','G','H','I','J','K','L','M','N']
SortId=2
points = {
    1:'E21',2:'E46',3:'E71',4:'E81',5:'E137',6:'E143',7:'E152'}

for i in range(1,3):
    totalPoint = 0
    eachPoints={}
    wb = load_workbook(f'excels/РЕЙТИНГ с изм.{i}.xlsx')
    currentSheet = wb.active
    kolledgeName = currentSheet['A4'].value
    Itogo = currentSheet['E153'].value
    ItogoVal = int("".join(filter(str.isdigit, Itogo)))
    kolledgeRating[kolledgeName] =  ItogoVal
    
    for point in points:
        kollegePoints = currentSheet[points[point]].value
        eachPoints[point]=kollegePoints

    for pointss in eachPoints:
        totalPoint = totalPoint + eachPoints[pointss]
    eachPoints['total'] = totalPoint
    kelledgeInfo[kolledgeName]= eachPoints

for i in sorted(kolledgeRating):
   kolledgeSortedRating[i]=kolledgeRating[i]
workbook = CreateWorkbook('Result.xlsx')
wb = load_workbook(workbook)
# wb.create_sheet('KolledgeResult')
currentSheet = wb.active
print(wb.active)
currentSheet['A1'].value = "Колледжи"
currentSheet['B1'].value = "Итоговые баллы"
currentSheet.auto_filter.ref = currentSheet.dimensions

for kolledge in kolledgeSortedRating:
    cellsA =f'A{SortId}'
    cellsB =f'B{SortId}'
    currentSheet[cellsA].value = kolledge
    currentSheet[cellsB].value = kolledgeSortedRating[kolledge]
    SortId= SortId+1


#Creating new Sheet and Sorting data
sheet = CreateSheet(workbook,'RazdelRating')
sheet = wb[sheet]
sheet['A1'].value = "Колледжи"
isFilter = True
index = 2
for keys in kelledgeInfo:
    j = 1
    for key in kelledgeInfo[keys]:
        abc = abat[j]
        val = str(kelledgeInfo[keys][key])
        FillRow(abc,index,key)
        if(isFilter):
            FillRow(abc,index-1,key)
        FillRow(abc,index,val)
        
        j=j+1
        sheet.auto_filter.ref = currentSheet.dimensions
    FillRow('A',index,keys)
    index += 1
    isFilter =False
        

wb.save(workbook)
print(kolledgeSortedRating,kelledgeInfo)

